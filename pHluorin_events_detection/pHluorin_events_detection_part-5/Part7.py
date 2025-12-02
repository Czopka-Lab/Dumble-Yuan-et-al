import openpyxl

# Open the Excel file
input_file_path = r'C:\x\Results\Halfwidthinput.xlsx'
workbook = openpyxl.load_workbook(input_file_path)
sheet = workbook.active

# Create a new workbook to store the interpolated data
output_file_path = r'C:\x\Results\Halfwidthcalc_interpol.xlsx'
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Iterate through all rows
for row in sheet.iter_rows(min_col=1, max_col=sheet.max_column):
    # Get the values in columns 1 to 4 (unchanged data)
    values_1_to_4 = [cell.value for cell in row[:4] if cell.value is not None]

    # Get the existing values in the row (columns 5 onwards) (edit from 4 to 5)
    values = [cell.value for cell in row[5:16] if cell.value is not None]

    # Check if all values in the row are numeric (skip non-numeric rows)
    if all(isinstance(value, (int, float)) for value in values):
        # Interpolate the values 10 times to the right
        interpolated_values = []
        for i in range(len(values) - 1):
            start_value = values[i]
            end_value = values[i + 1]
            step = (end_value - start_value) / 10
            interpolated_values.extend([start_value + step * j for j in range(11)])

        # Write the values from columns 1 to 4 followed by interpolated data to the new workbook
        output_sheet.append(values_1_to_4 + interpolated_values)

# Save the new Excel file with interpolated data
output_workbook.save(output_file_path)

# Close both workbooks
workbook.close()
output_workbook.close()

print(f"Interpolated data saved to {output_file_path}")
